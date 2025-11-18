import os
import sys
import shutil
import pandas as pd
from pathlib import Path
import datetime
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Tuple, Optional
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------- 공용 유틸 ----------------------
# 이사진 성명
director_list = ['강두영', '강병수', '정재황']

def director_check(name: str) -> int:
    """
    이사직은 철야근무가 있으므로 7줄, 그 외는 6줄을 조정함
    """
    row_adj = 7
    if name in director_list:
        row_adj = 8
    return row_adj


def resource_path(relative_path: str) -> str:
    """
    PyInstaller 빌드 후에도 템플릿 등 리소스 접근 가능
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def safe_to_hm(value: str) -> Optional[str]:
    """
    시간값을 'HH:MM' 문자열로 변환
    """
    if pd.isna(value):
        return None
    try:
        t = pd.to_datetime(value)
        return t.strftime("%H:%M")
    except Exception:
        return None


# ---------------------- 데이터 처리 ----------------------
def process_factory(factory_path: str):
    """
    공장 근태 raw 데이터 파싱
    """
    if not factory_path:
        return pd.DataFrame(columns=['성명','사번','부서','날짜','출근시간','퇴근시간'])
    try:
        for i in range(5):
            df_tmp = pd.read_excel(factory_path, engine='openpyxl', header=i)
            cols = [str(c).strip() for c in df_tmp.columns]
            if "출입날짜" in cols and "출입시간" in cols:
                df = df_tmp
                break
        else:
            raise ValueError("출입날짜 / 출입시간 컬럼을 찾을 수 없습니다.")
    except Exception as e:
        messagebox.showerror("파일 읽기 실패", f"공장 파일을 읽는 중 에러:\n{e}")
        return pd.DataFrame(columns=['성명','사번','부서','날짜','출근시간','퇴근시간'])

    # 모든 컬럼명에서 공백 제거
    df.columns = df.columns.str.strip()

    rename_map = {
        "출입날짜": "날짜",
        "출입시간": "시간",
        "사  번": "사번",
        "이  름": "성명",
        "기능키": "기능키"
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    df["시간_raw"] = df["시간"].astype(str)
    df["시간_hm"] = df["시간_raw"].apply(safe_to_hm)
    df["날짜"] = pd.to_datetime(df["날짜"]).dt.date

    result = df.groupby(["성명", "사번", "날짜"], as_index=False).agg(
        출근시간=("시간_hm", "min"),
        퇴근시간=("시간_hm", "max")
    )
    result["구분"] = "공장"
    return result


def process_office(office_path: str, year: str, month: str):
    """
    사무실 근태 raw 데이터 파싱
    """
    if not office_path:
        return pd.DataFrame(columns=['성명','사번','부서','날짜','출근시간','퇴근시간'])
    try:
        wb = load_workbook(office_path, data_only=True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("파일 읽기 실패", f"사무실 파일을 읽는 중 에러:\n{e}")
        return pd.DataFrame(columns=['성명','사번','부서','날짜','출근시간','퇴근시간'])

    data = []
    for r in range(5, ws.max_row+1, 3):  # 실제 오프셋은 샘플에서 조정 필요
        name = ws.cell(r, 11).value
        emp_num = ws.cell(r, 3).value
        dept = ws.cell(r, 21).value
        if not name: continue
        for c in range(5, 36):
            time_stamp = None
            time_stamp_split = None
            start = None
            end = None
            date = c-4

            time_stamp = ws.cell(r+1, date).value
            if time_stamp is not None:
                time_stamp_split = time_stamp.splitlines()
                start = time_stamp_split[0]
                end = time_stamp_split[-1]

            if start or end:
                data.append([name, emp_num, dept, date, start, end])
    df = pd.DataFrame(data, columns=['성명','사번','부서','일','출근시간','퇴근시간'])

    # 모든 컬럼명에서 공백 제거
    df.columns = df.columns.str.strip()

    df['사번'] = df['사번'].astype(int)
    df['날짜'] = pd.to_datetime(f"2025-10-" + df['일'].astype(str)).dt.date
    df = df.drop(columns=['부서','일'])
    df['구분'] = '사무실'
    df = df.reindex(columns=['성명','사번','날짜','출근시간','퇴근시간','구분'])
    return df


def merge_table(factory_df, office_df):
    """
    두 시트 데이터 결합
    """
    df = pd.concat([factory_df, office_df], ignore_index=True)

    # 성명, 사번, 날짜 기준으로 그룹화하여 병합
    df_summary = df.groupby(['성명', '사번', '날짜'], as_index=False).agg({
        '출근시간': 'min',   # 가장 빠른 출근시간
        '퇴근시간': 'max',   # 가장 늦은 퇴근시간
        '구분': lambda x: ','.join(sorted(set(x.dropna())))  # 여러 구분값이 있을 경우 병합
    })

    return df_summary


# ---------------------- 보고용 시트 매핑 ----------------------
def to_time(value) -> Optional[datetime.time]:
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.datetime.strptime(value.strip(), "%H:%M").time()
        except ValueError:
            pass
    return None


def to_day_on_sheet(day: int, day_of_weekday: int):
    day_of_week_list = ['월', '화', '수', '목', '금', '토', '일']
    result_str = str(day) + '(' + day_of_week_list[day_of_weekday] + ')'
    return result_str


def iter_parse_report_sheet(ws: Worksheet, year_str: str, month_str: str) -> Dict[str, Dict[int, Tuple[int, int, int]]]:
    result_cells = {}
    for r in range(2, ws.max_row + 1):
        day_cells = {}

        val = ws.cell(r+1, 1).value
        if val != '성명' and pd.notna(val):
            val_split = val.splitlines()
            name = val_split[0]
            row_adj = director_check(name)

            for c in range(3, ws.max_column + 1):
                day = ws.cell(r, c).value
                if pd.notna(day) and isinstance(day, int):
                    if 1 <= int(day) <= 31:
                        day_of_week = datetime.date(int(year_str), int(month_str), day).weekday()
                        day_cells[day] = (r, c, day_of_week)
                        # 보고용 시트에 날짜 입력
                        ws.cell(r, c).value = to_day_on_sheet(day, day_of_week)

            for c in range(3, ws.max_column + 1):
                day = ws.cell(r+row_adj, c).value
                if pd.notna(day) and isinstance(day, int):
                    if 1 <= int(day) <= 31:
                        day_of_week = datetime.date(int(year_str), int(month_str), day).weekday()
                        day_cells[day] = (r+row_adj, c, day_of_week)
                        # 보고용 시트에 날짜 입력
                        ws.cell(r+row_adj, c).value = to_day_on_sheet(day, day_of_week)

            result_cells[name] = day_cells
    return result_cells


def find_target_row(ws: Worksheet, name: str, header_row: int) -> Optional[int]:
    for r in range(header_row + 1, min(ws.max_row, header_row + 6)):
        n_name = ''
        n = ws.cell(r, 1).value
        if n is not None:
            n_split = n.splitlines()
            n_name = n_split[0]
        else:
            row_adj = director_check(name)
            n_adj = ws.cell(r-row_adj, 1).value
            n_split = n_adj.splitlines()
            n_name = n_split[0]
        if n_name == name:
            return r
    return None


def apply_attendance(ws: Worksheet, df: pd.DataFrame, year_str: str, month_str: str):
    """
    정리테이블(df)의 근태 데이터를 보고용 시트(ws)에 반영
    - 출근, 퇴근, 근무시간, 연장근무, 휴일근무 자동 기입
    """
    result_cells = iter_parse_report_sheet(ws, year_str, month_str)
    if not (result_cells):
        return

    for key in result_cells:
        emp_records = df[(df["성명"] == key)]
        if emp_records.empty:
            continue

        for _, record in emp_records.iterrows():
            if pd.isna(record["날짜"]):
                continue
            date = pd.to_datetime(record["날짜"])
            day = int(date.day)

            day_cells = result_cells[key]
            if day not in day_cells:
                continue

            s, e = to_time(record["출근시간"]), to_time(record["퇴근시간"])
            if not s and not e:
                continue

            header_row, col, day_of_weekday = day_cells[day]

            # 이름이 있는 행(n행)을 찾고, 그 아래 5행 블록을 채운다.
            target_row = find_target_row(ws, key, header_row)
            if not target_row:
                continue

            # 이사직은 철야근무가 있으므로 7줄, 그 외는 6줄을 조정함
            director_flag = False
            if key in director_list:
                director_flag = True

            row_in = target_row             # 출근시간
            row_out = target_row + 1        # 퇴근시간
            row_work = target_row + 2       # 근무시간
            row_ot = target_row + 3         # 연장근무
            row_holiday = target_row + 4    # 휴일근무
            if director_flag:
                row_night = target_row + 5  # 철야근무

            # 출근시간 및 퇴근시간
            if s:
                ws.cell(row_in, col).value = s.strftime("%H:%M")
            if e:
                ws.cell(row_out, col).value = e.strftime("%H:%M")

            # 근무시간 계산
            work_hours = None
            if s and e:
                s_dt = datetime.datetime.combine(datetime.datetime.today(), s)
                e_dt = datetime.datetime.combine(datetime.datetime.today(), e)
                work_hours = round((e_dt - s_dt).seconds / 3600, 1)
                ws.cell(row_work, col).value = work_hours

            # 연장근무 (출근 < 8:00, 퇴근 > 17:00)
            if s and e and (s > datetime.time(5, 0) and s < datetime.time(8, 0)) or (e > datetime.time(17, 0) and e < datetime.time(23, 0)):
                before8 = 0
                after17 = 0

                if s > datetime.time(5, 0) and s < datetime.time(8, 0):
                    before8 = (datetime.datetime.combine(datetime.datetime.today(), datetime.time(8, 0)) -
                                datetime.datetime.combine(datetime.datetime.today(), s)).seconds / 3600
                if e > datetime.time(17, 0) and e < datetime.time(23, 0):
                    after17 = (datetime.datetime.combine(datetime.datetime.today(), e) -
                                datetime.datetime.combine(datetime.datetime.today(), datetime.time(17, 0))).seconds / 3600

                ws.cell(row_ot, col).value = round(before8 + after17, 1)
            else:
                ws.cell(row_ot, col).value = "-"

            # 휴일근무 (토:5, 일:6)
            if day_of_weekday in (5, 6) and work_hours:
                ws.cell(row_holiday, col).value = work_hours
            else:
                ws.cell(row_holiday, col).value = "-"

            # 철야근무 (근무 시간 -> 23:00 - 05:00)
            if director_flag:
                if s and e and (e > datetime.time(23, 0) and s < datetime.time(5, 0)):
                    before5 = 0
                    after23 = 0

                    if s < datetime.time(5, 0):
                        before5 = (datetime.datetime.combine(datetime.datetime.today(), datetime.time(5, 0)) -
                                    datetime.datetime.combine(datetime.datetime.today(), s)).seconds / 3600
                    if e > datetime.time(23, 0):
                        after23 = (datetime.datetime.combine(datetime.datetime.today(), e) -
                                    datetime.datetime.combine(datetime.datetime.today(), datetime.time(23, 0))).seconds / 3600

                    ws.cell(row_night, col).value = round(before5 + after23, 1)
                else:
                    ws.cell(row_night, col).value = "-"


def update_report_sheet(wb, year_str: str, month_str: str):
    """
    정리테이블 → {YY.MM} 시트 반영
    """
    sheet_name = f"{year_str[2:]}.{month_str}"
    if "정리테이블" not in wb.sheetnames:
        return
    ws_data = wb["정리테이블"]
    df = pd.DataFrame(list(ws_data.values)[1:], columns=list(ws_data.values)[0])
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    tmp_sheet = wb['tmp']
    report_sheet = wb.copy_worksheet(tmp_sheet)
    report_sheet.title = sheet_name
    report_sheet['A1'] = '㈜대영인텍 출퇴근기록부 - ' + str(year_str) + '년 ' + str(month_str) + '월'
    apply_attendance(wb[sheet_name], df, year_str, month_str)

    wb.move_sheet('tmp', 4)
    wb.move_sheet(sheet_name, -3)
    wb.move_sheet('정리테이블', -2)


# ---------------------- 엑셀 갱신 ----------------------
def update_excel(factory_df, office_df, merged_df, year_str, month_str):
    out_path = Path(f"output/근태기록_{year_str}_{month_str}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    template = Path(resource_path("template.xlsx"))

    if not out_path.exists():
        shutil.copy(template, out_path)

    wb = load_workbook(out_path)

    def update_sheet(sheet_name, df):
        from openpyxl.utils.dataframe import dataframe_to_rows
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
            ws.delete_rows(2, ws.max_row)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    if factory_df is not None:
        update_sheet("공장데이터", factory_df)
    if office_df is not None:
        update_sheet("사무실데이터", office_df)
    if merged_df is not None:
        update_sheet("정리테이블", merged_df)

    update_report_sheet(wb, year_str, month_str)
    wb.save(out_path)
    return out_path


# ---------------------- GUI ----------------------
VERSION = "v1.1.1"
BUILD_DATE = "2025-11-18"
DEVELOPER = "왕형순"


def show_version_info():
    msg = f"근태 정리 자동화 프로그램\n\n버   전: {VERSION}\n빌드일: {BUILD_DATE}\n개발자: {DEVELOPER}"
    messagebox.showinfo("버전 정보", msg)


def run_processing(progress_var, execute_button):
    """
    비동기 실행 함수
    """
    try:
        # Step 1: 입력값 확인
        # 진행률 표시 (실행중 모드 진입)
        root.geometry("480x250")
        progress_frame.grid(row=6, column=0, columnspan=3, pady=10)
        progress_var.set(0)

        execute_button.config(text="실행중...", state=tk.DISABLED)

        # 단계별 진행률 갱신
        progress_var.set(10)
        root.update()

        year_str = year_entry.get().strip()
        month_str = month_entry.get().strip()
        factory_path = factory_entry.get().strip()
        office_path = office_entry.get().strip()

        if not (year_str and month_str):
            messagebox.showerror("입력 오류", "년, 월을 모두 입력하세요.")
            return

        # Step 2: 데이터 처리
        progress_var.set(30)
        root.update()

        factory_df = process_factory(factory_path) if factory_path else None
        office_df = process_office(office_path, year_str, month_str) if office_path else None
        if factory_df is not None and office_df is not None:
            merged_df = merge_table(factory_df, office_df)

        # Step 3: 엑셀 갱신
        progress_var.set(70)
        root.update()

        if merged_df is not None:
            out_path = update_excel(factory_df, office_df, merged_df, year_str, month_str)
            progress_var.set(100)
            root.update()
            messagebox.showinfo("완료", f"근태기록 갱신 완료!\n\n{out_path}")
        else:
            messagebox.showwarning("데이터 없음", "입력된 근태 데이터가 없습니다.")

    except Exception as e:
        messagebox.showerror("오류", f"실행 중 오류 발생:\n{e}")

    finally:
        # 실행 후 복원
        execute_button.config(text="실행", state=tk.NORMAL)
        progress_var.set(0)
        progress_frame.grid_remove()
        root.geometry("480x200")
        root.update()


def execute_async():
    """
    실행 버튼 클릭 시 비동기로 실행
    """
    thread = threading.Thread(target=run_processing, args=(progress_var, execute_button))
    thread.start()


# ---------------------- GUI 생성 ----------------------
root = tk.Tk()
root.title("근태기록 자동 정리기")
root.geometry("480x200")

tk.Label(root, text="년도 (YYYY):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
year_entry = tk.Entry(root, width=40); year_entry.grid(row=0, column=1)

tk.Label(root, text="월 (MM):").grid(row=1, column=0, sticky="e", padx=5, pady=5)
month_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
month_entry = ttk.Combobox(root, values=month_list, width=37); month_entry.grid(row=1, column=1)

tk.Label(root, text="공장 근태 파일:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
factory_entry = tk.Entry(root, width=40); factory_entry.grid(row=2, column=1)
tk.Button(root, text="찾기", width=8, command=lambda: factory_entry.insert(0, filedialog.askopenfilename())).grid(row=2, column=2, sticky="e")

tk.Label(root, text="사무실 근태 파일:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
office_entry = tk.Entry(root, width=40); office_entry.grid(row=3, column=1)
tk.Button(root, text="찾기", width=8, command=lambda: office_entry.insert(0, filedialog.askopenfilename())).grid(row=3, column=2, sticky="e")

# 실행 버튼 + 버전 버튼
execute_button = tk.Button(root, text="실행", width=15, command=execute_async)
execute_button.grid(row=5, column=1, pady=15)
tk.Button(root, text="버전정보", width=10, command=show_version_info).grid(row=5, column=2, pady=15)

# 진행률 표시 프레임 (기본 숨김)
progress_frame = tk.Frame(root)
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(progress_frame, variable=progress_var, maximum=100, length=300)
progress_bar.grid(row=0, column=0, padx=10)
progress_frame.grid_remove()  # 처음엔 숨김

root.mainloop()
